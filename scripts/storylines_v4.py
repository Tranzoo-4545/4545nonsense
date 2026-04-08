#!/usr/bin/env python3
"""
Lichess4545 Storylines Generator v4

Calculates "hype scores" for upcoming pairings to find the most compelling matchups.

v4 Changes:
- Added Titled Player factor (+1.5 per titled player)
- Added 100th Win milestone (+2.5 for "going for 100th win")
- Tiered Return Match (5-9 seasons +2, 10-19 +3, 20+ +4)
- League Champions bonus (+1 if former teammates won league together)
- Added ICS calendar support for scheduled game times

v2 Changes:
- Added Excel export option
- Added Vets First Meeting factor
- Auto-refreshes season data (no need to run enricher first)

Usage:
    py storylines_v4.py --season 47 --round 5
    py storylines_v4.py --season 47              # Auto-detect next round
    py storylines_v4.py --season 47 --ics calendar.ics  # Include scheduled times
    py storylines_v4.py --season 47 --excel storylines.xlsx
"""

import argparse
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from collections import defaultdict

# Import standings calculator
try:
    from standings import calculate_standings
except ImportError:
    print("Error: standings.py must be in the same directory")
    sys.exit(1)

CACHE_DIR = Path.home() / '.lichess4545_cache' / 'retention'
TITLES_CACHE_FILE = Path.home() / '.lichess4545_cache' / 'player_titles.json'
LEAGUE_API_URL = "https://www.lichess4545.com/api/get_season_games/?league=team4545&include_unplayed=true&season={}"
LICHESS_USERS_URL = "https://lichess.org/api/users"
ICS_URL_TEMPLATE = "https://www.lichess4545.com/team4545/season/{}/pairings/calendar.ics"


def fetch_season_from_api(season):
    """Fetch season data from Lichess4545 API."""
    try:
        import requests
    except ImportError:
        print("Error: requests library required. Run: pip install requests")
        sys.exit(1)
    
    url = LEAGUE_API_URL.format(season)
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return response.json()


def save_season_to_cache(season, data):
    """Save season data to cache."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_file = CACHE_DIR / f"season_{season}_with_unplayed.json"
    with open(cache_file, 'w', encoding='utf-8') as f:
        json.dump(data, f)


def refresh_season_cache(season, force=False):
    """Refresh season data from league API if needed."""
    cache_file = CACHE_DIR / f"season_{season}_with_unplayed.json"
    
    # Check if cache exists and is recent (less than 1 hour old)
    if cache_file.exists() and not force:
        age_seconds = time.time() - cache_file.stat().st_mtime
        if age_seconds < 3600:  # 1 hour
            return False  # No refresh needed
    
    print(f"Refreshing season {season} data from league API...")
    try:
        data = fetch_season_from_api(season)
        save_season_to_cache(season, data)
        print(f"  Fetched {len(data.get('games', []))} games")
        return True
    except Exception as e:
        print(f"  Warning: Could not refresh season {season}: {e}")
        return False


def load_season(season, auto_fetch=True):
    """Load season data from cache, fetching if needed."""
    cache_file = CACHE_DIR / f"season_{season}_with_unplayed.json"
    
    if cache_file.exists():
        with open(cache_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    if auto_fetch:
        print(f"Season {season} not cached, fetching from API...")
        try:
            data = fetch_season_from_api(season)
            save_season_to_cache(season, data)
            time.sleep(0.5)  # Be nice to API
            return data
        except Exception as e:
            print(f"  Error fetching season {season}: {e}")
            return None
    
    print(f"Season {season} not cached. Run: py achievements.py --max {season}")
    return None


def load_all_seasons(max_season):
    """Load all seasons for historical data."""
    all_data = {}
    for s in range(1, max_season + 1):
        data = load_season(s)
        if data:
            all_data[s] = data
    return all_data


def get_championship_teams(all_season_data):
    """Get the winning team for each completed season."""
    champions = {}  # season -> team_name
    
    for season_num, season_data in all_season_data.items():
        # Only consider completed seasons (8 rounds)
        rounds = set(g.get('round', 0) for g in season_data.get('games', []) if g.get('game_id'))
        if len(rounds) < 8:
            continue  # Season not complete
        
        standings = calculate_standings(season_data)
        if standings:
            # First place team (standings use 'name' key)
            champions[season_num] = standings[0]['name']
    
    return champions


def load_player_titles():
    """Load cached player titles."""
    if TITLES_CACHE_FILE.exists():
        try:
            with open(TITLES_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}


def save_player_titles(titles):
    """Save player titles to cache."""
    TITLES_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(TITLES_CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(titles, f)


def fetch_player_titles(usernames):
    """Fetch player titles from Lichess API (bulk request)."""
    try:
        import requests
    except ImportError:
        return {}
    
    if not usernames:
        return {}
    
    # Load existing cache
    cached = load_player_titles()
    
    # Filter to usernames we don't have cached
    usernames_lower = [u.lower() for u in usernames]
    to_fetch = [u for u in usernames_lower if u not in cached]
    
    if not to_fetch:
        return {u: cached.get(u) for u in usernames_lower}
    
    print(f"  Fetching titles for {len(to_fetch)} players from Lichess...")
    
    # Lichess API accepts up to 300 users per request
    titles = dict(cached)  # Start with cached data
    
    for i in range(0, len(to_fetch), 300):
        batch = to_fetch[i:i+300]
        try:
            response = requests.post(
                LICHESS_USERS_URL,
                data=','.join(batch),
                headers={'Accept': 'application/x-ndjson'},
                timeout=30
            )
            
            if response.status_code == 200:
                content = response.text.strip()
                
                # Handle both nd-json (one object per line) and JSON array formats
                for line in content.split('\n'):
                    if not line:
                        continue
                    try:
                        parsed = json.loads(line)
                        
                        # If it's a list (JSON array), iterate through it
                        if isinstance(parsed, list):
                            for user in parsed:
                                if isinstance(user, dict):
                                    username = user.get('id', '').lower()
                                    title = user.get('title')
                                    titles[username] = title
                        # If it's a dict (single user object from nd-json)
                        elif isinstance(parsed, dict):
                            username = parsed.get('id', '').lower()
                            title = parsed.get('title')
                            titles[username] = title
                    except json.JSONDecodeError:
                        continue
            
            time.sleep(0.5)  # Be nice to API
        except Exception as e:
            print(f"    Warning: Error fetching titles: {e}")
    
    # Cache all results (including None for non-titled)
    for u in to_fetch:
        if u not in titles:
            titles[u] = None
    
    save_player_titles(titles)
    
    return {u: titles.get(u) for u in usernames_lower}


def build_player_history(all_season_data, current_season=None, before_round=None, championship_teams=None):
    """Build player historical data from all seasons.
    
    For current_season, only counts games before before_round.
    championship_teams: dict of season -> winning team name
    """
    players = defaultdict(lambda: {
        'total_games': 0,
        'total_wins': 0,
        'opponents': defaultdict(lambda: {'games': 0, 'wins': 0, 'losses': 0, 'draws': 0}),
        'teams': defaultdict(int),  # team_name -> games played
        'championship_teams': set(),  # teams they won the league with (4+ games)
        'seasons_played': set(),  # set of season numbers
        'games_per_season': defaultdict(int),  # season -> game count
    })
    
    if championship_teams is None:
        championship_teams = {}
    
    for season_num, season_data in all_season_data.items():
        for game in season_data.get('games', []):
            white = (game.get('white') or '').lower()
            black = (game.get('black') or '').lower()
            result = game.get('result', '')
            game_id = game.get('game_id')
            round_num = game.get('round', 0)
            white_team = game.get('white_team') or game.get('white_team_name', '')
            black_team = game.get('black_team') or game.get('black_team_name', '')
            
            if not white or not black or not game_id:
                continue
            
            # For current season, only count games before target round
            if current_season and season_num == current_season and before_round:
                if round_num >= before_round:
                    continue
            
            # Count played games
            players[white]['total_games'] += 1
            players[black]['total_games'] += 1
            
            # Track teams played for
            if white_team:
                players[white]['teams'][white_team] += 1
            if black_team:
                players[black]['teams'][black_team] += 1
            
            # Track seasons played
            players[white]['seasons_played'].add(season_num)
            players[black]['seasons_played'].add(season_num)
            players[white]['games_per_season'][season_num] += 1
            players[black]['games_per_season'][season_num] += 1
            
            # Track head-to-head
            players[white]['opponents'][black]['games'] += 1
            players[black]['opponents'][white]['games'] += 1
            
            if result in ('1-0', '1X-0F'):
                players[white]['opponents'][black]['wins'] += 1
                players[black]['opponents'][white]['losses'] += 1
                players[white]['total_wins'] += 1
            elif result in ('0-1', '0F-1X'):
                players[black]['opponents'][white]['wins'] += 1
                players[white]['opponents'][black]['losses'] += 1
                players[black]['total_wins'] += 1
            elif result in ('1/2-1/2', '1/2Z-1/2Z'):
                players[white]['opponents'][black]['draws'] += 1
                players[black]['opponents'][white]['draws'] += 1
    
    # Now calculate championship teams for each player
    for player, data in players.items():
        for team, games in data['teams'].items():
            if games >= 4:  # Only count if played 4+ games for team
                # Check if this team won any season
                for season, champ_team in championship_teams.items():
                    if team == champ_team and season in data['seasons_played']:
                        # Verify they played for this team in this season
                        # (this is approximate - checks if they played for the team ever)
                        data['championship_teams'].add(team)
    
    return players


def get_career_streaks(all_season_data, current_season, before_round):
    """Calculate career-wide win/loss streaks for each player."""
    player_games = defaultdict(list)
    
    for season_num, season_data in all_season_data.items():
        for game in season_data.get('games', []):
            white = (game.get('white') or '').lower()
            black = (game.get('black') or '').lower()
            result = game.get('result', '')
            game_id = game.get('game_id')
            round_num = game.get('round', 0)
            
            # Only count actual played games (with game_id)
            if not white or not black or not game_id:
                continue
            
            # For current season, only count games before target round
            if season_num == current_season and round_num >= before_round:
                continue
            
            # Determine results for each player
            if result in ('1-0', '1X-0F'):
                player_games[white].append((season_num, round_num, 'W'))
                player_games[black].append((season_num, round_num, 'L'))
            elif result in ('0-1', '0F-1X'):
                player_games[white].append((season_num, round_num, 'L'))
                player_games[black].append((season_num, round_num, 'W'))
            elif result in ('1/2-1/2', '1/2Z-1/2Z'):
                player_games[white].append((season_num, round_num, 'D'))
                player_games[black].append((season_num, round_num, 'D'))
    
    # Calculate current streak for each player
    streaks = {}
    for player, games in player_games.items():
        games.sort(key=lambda x: (x[0], x[1]))  # Sort by season, then round
        
        if not games:
            streaks[player] = {'type': None, 'length': 0}
            continue
        
        # Get streak from most recent games
        last_result = games[-1][2]  # (season, round, result)
        if last_result == 'D':
            streaks[player] = {'type': 'draw', 'length': 1}
            continue
        
        streak_type = 'win' if last_result == 'W' else 'loss'
        streak_len = 0
        
        for season, round_num, result in reversed(games):
            if (streak_type == 'win' and result == 'W') or (streak_type == 'loss' and result == 'L'):
                streak_len += 1
            else:
                break
        
        streaks[player] = {'type': streak_type, 'length': streak_len}
    
    return streaks


def get_round_pairings(season_data, target_round=None):
    """Get all pairings for a specific round (played or unplayed)."""
    games = season_data.get('games', [])
    
    # Group all games by round (normalize to int)
    rounds = defaultdict(list)
    for g in games:
        r = g.get('round')
        if r is not None:
            # Normalize to int for consistent comparison
            try:
                r_int = int(r)
                rounds[r_int].append(g)
            except (ValueError, TypeError):
                pass
    
    if target_round:
        target_round = int(target_round)  # Ensure target is also int
        if target_round in rounds:
            return target_round, rounds[target_round]
        else:
            print(f"No games found for round {target_round}")
            return None, []
    
    # Auto-detect: find earliest round with any unplayed games
    for r in sorted(rounds.keys()):
        has_unplayed = any(not g.get('game_id') for g in rounds[r])
        if has_unplayed:
            return r, rounds[r]
    
    # If all games played, return the last round
    if rounds:
        last_round = max(rounds.keys())
        return last_round, rounds[last_round]
    
    return None, []


# =============================================================================
# ICS CALENDAR PARSING
# =============================================================================

def parse_ics_file(filepath):
    """Parse ICS file and extract scheduled game times."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    return parse_ics_content(content)


def fetch_ics_calendar(season):
    """Fetch ICS calendar from league website."""
    try:
        import requests
    except ImportError:
        print("Error: requests library required for ICS fetch")
        return []
    
    url = ICS_URL_TEMPLATE.format(season)
    print(f"Fetching ICS from {url}...")
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return parse_ics_content(response.text)


def parse_ics_content(content):
    """Parse ICS content and extract scheduled game times."""
    content = content.replace('\r\n ', '').replace('\n ', '')
    
    events = []
    for block in content.split('BEGIN:VEVENT'):
        if 'END:VEVENT' not in block:
            continue
        
        summary_match = re.search(r'SUMMARY:(.+?)(?:\r?\n|$)', block)
        dtstart_match = re.search(r'DTSTART:(\d{8}T\d{6}Z?)', block)
        
        if not summary_match or not dtstart_match:
            continue
            
        summary = summary_match.group(1).strip()
        # Unescape ICS special characters
        summary = summary.replace('\\,', ',').replace('\\;', ';').replace('\\\\', '\\')
        dtstart_str = dtstart_match.group(1)
        
        if dtstart_str.endswith('Z'):
            scheduled_time = datetime.strptime(dtstart_str, '%Y%m%dT%H%M%SZ')
        else:
            scheduled_time = datetime.strptime(dtstart_str, '%Y%m%dT%H%M%S')
        
        match = re.match(r'(.+?)\s*\((.+?)\)\s*vs\s*(.+?)\s*\((.+?)\)', summary)
        if match:
            p1, t1, p2, t2 = match.groups()
            events.append({
                'white': p1.strip().lower(),
                'white_team': t1.strip(),
                'black': p2.strip().lower(),
                'black_team': t2.strip(),
                'scheduled_time': scheduled_time,
            })
    
    return events


def build_schedule_lookup(ics_events):
    """Build a lookup dict from player pairs to scheduled times."""
    lookup = {}
    for e in ics_events:
        player_key = tuple(sorted([e['white'], e['black']]))
        lookup[player_key] = e['scheduled_time']
    return lookup


def get_standings_multiplier(round_num):
    """Get multiplier for standings-related factors based on round number."""
    multipliers = {
        1: 0.0,
        2: 0.25,
        3: 0.5,
        4: 0.75,
        5: 1.0,
        6: 1.0,
        7: 1.5,
        8: 1.5,
    }
    return multipliers.get(round_num, 1.0)


def calculate_hype_score(pairing, player_history, standings, streaks, team_battle_ranks, current_season, player_titles=None, round_num=None):
    """Calculate hype score for a pairing. Returns (score, reasons, breakdown)."""
    white = (pairing.get('white') or '').lower()
    black = (pairing.get('black') or '').lower()
    white_team = pairing.get('white_team') or pairing.get('white_team_name', '')
    black_team = pairing.get('black_team') or pairing.get('black_team_name', '')
    
    if player_titles is None:
        player_titles = {}
    
    score = 0
    reasons = []
    breakdown = {
        'milestone': 0,
        'win_milestone': 0,
        'h2h': 0,
        'form_clash': 0,
        'team_battle': 0,
        'veteran': 0,
        'return_match': 0,
        'former_teammates': 0,
        'vets_first_meeting': 0,
        'debutants_duel': 0,
        'titled': 0,
    }
    
    white_data = player_history.get(white, {'total_games': 0, 'total_wins': 0, 'opponents': {}, 'championship_teams': set()})
    black_data = player_history.get(black, {'total_games': 0, 'total_wins': 0, 'opponents': {}, 'championship_teams': set()})
    
    # TITLED PLAYER (0-3 points) - +1.5 per titled player
    white_title = player_titles.get(white)
    black_title = player_titles.get(black)
    
    if white_title and black_title:
        pts = 3
        score += pts
        breakdown['titled'] = pts
        reasons.append(f"👑 Titled showdown! {white_title} vs {black_title}")
    elif white_title:
        pts = 1.5
        score += pts
        breakdown['titled'] = pts
        reasons.append(f"🎖️ {white_title} {white} in the mix")
    elif black_title:
        pts = 1.5
        score += pts
        breakdown['titled'] = pts
        reasons.append(f"🎖️ {black_title} {black} in the mix")
    
    # MILESTONE - GAMES (0-6 points)
    for player, data, color in [(white, white_data, 'White'), (black, black_data, 'Black')]:
        next_game = data.get('total_games', 0) + 1
        if next_game == 300:
            pts = 3
            score += pts
            breakdown['milestone'] += pts
            reasons.append(f"🎂 {player}'s 300th game!")
        elif next_game == 200:
            pts = 3
            score += pts
            breakdown['milestone'] += pts
            reasons.append(f"🎂 {player}'s 200th game!")
        elif next_game == 100:
            pts = 3
            score += pts
            breakdown['milestone'] += pts
            reasons.append(f"💯 {player}'s 100th game!")
        elif next_game == 45:
            pts = 2
            score += pts
            breakdown['milestone'] += pts
            reasons.append(f"🎯 {player}'s 45th game!")
    
    # MILESTONE - 100TH WIN (0-2.5 points)
    for player, data in [(white, white_data), (black, black_data)]:
        wins = data.get('total_wins', 0)
        if wins == 99:
            pts = 2.5
            score += pts
            breakdown['win_milestone'] += pts
            reasons.append(f"🏆 {player} going for 100th win!")
    
    # HEAD-TO-HEAD (0-4 points)
    h2h = white_data['opponents'].get(black, {'games': 0})
    h2h_games = h2h.get('games', 0)
    
    if h2h_games >= 3:
        pts = 4
        score += pts
        breakdown['h2h'] = pts
        w_wins = h2h.get('wins', 0)
        w_losses = h2h.get('losses', 0)
        w_draws = h2h.get('draws', 0)
        reasons.append(f"⚔️ Rivals! Met {h2h_games}x before ({w_wins}-{w_losses}-{w_draws})")
    elif h2h_games == 2:
        pts = 3
        score += pts
        breakdown['h2h'] = pts
        w_wins = h2h.get('wins', 0)
        w_losses = h2h.get('losses', 0)
        w_draws = h2h.get('draws', 0)
        reasons.append(f"🔄 Rematch #{h2h_games + 1} ({w_wins}-{w_losses}-{w_draws})")
    elif h2h_games == 1:
        pts = 2
        score += pts
        breakdown['h2h'] = pts
        w_wins = h2h.get('wins', 0)
        w_losses = h2h.get('losses', 0)
        w_draws = h2h.get('draws', 0)
        reasons.append(f"🔄 Rematch ({w_wins}-{w_losses}-{w_draws})")
    
    # FORM CLASH (0-3 points)
    white_streak = streaks.get(white, {'type': None, 'length': 0})
    black_streak = streaks.get(black, {'type': None, 'length': 0})
    
    # Hot streak vs hot streak
    if white_streak['type'] == 'win' and black_streak['type'] == 'win':
        if white_streak['length'] >= 3 and black_streak['length'] >= 3:
            pts = 3
            score += pts
            breakdown['form_clash'] = pts
            reasons.append(f"🔥 Hot streak clash! ({white_streak['length']}W vs {black_streak['length']}W)")
        elif white_streak['length'] >= 2 and black_streak['length'] >= 2:
            pts = 2
            score += pts
            breakdown['form_clash'] = pts
            reasons.append(f"🔥 Both on win streaks ({white_streak['length']}W vs {black_streak['length']}W)")
    
    # Cold streak vs cold streak
    elif white_streak['type'] == 'loss' and black_streak['type'] == 'loss':
        if white_streak['length'] >= 3 and black_streak['length'] >= 3:
            pts = 3
            score += pts
            breakdown['form_clash'] = pts
            reasons.append(f"❄️ Desperation match! ({white_streak['length']}L vs {black_streak['length']}L)")
        elif white_streak['length'] >= 2 and black_streak['length'] >= 2:
            pts = 2
            score += pts
            breakdown['form_clash'] = pts
            reasons.append(f"❄️ Both struggling ({white_streak['length']}L vs {black_streak['length']}L)")
    
    # TEAM BATTLE RANK (0-4 points, scaled by round)
    match_key = tuple(sorted([white_team, black_team]))
    if match_key in team_battle_ranks:
        rank = team_battle_ranks[match_key]
        standings_mult = get_standings_multiplier(round_num) if round_num else 1.0
        
        if rank == 1:
            base_pts = 4
            pts = base_pts * standings_mult
            if pts > 0:
                score += pts
                breakdown['team_battle'] = pts
                reasons.append(f"🏆 Huge for team standings!")
        elif rank == 2:
            base_pts = 3
            pts = base_pts * standings_mult
            if pts > 0:
                score += pts
                breakdown['team_battle'] = pts
                reasons.append(f"🥈 Big for team standings!")
        elif rank == 3:
            base_pts = 2
            pts = base_pts * standings_mult
            if pts > 0:
                score += pts
                breakdown['team_battle'] = pts
                reasons.append(f"🥉 Important for team standings!")
    
    # VETERAN (0-4 points)
    vet_score = 0
    for player, data in [(white, white_data), (black, black_data)]:
        games = data['total_games']
        player_vet = min(games / 100, 2)
        vet_score += player_vet
    
    if vet_score >= 2:
        score += vet_score
        breakdown['veteran'] = vet_score
        reasons.append(f"🎖️ Veterans ({white_data['total_games']}+{black_data['total_games']} games)")
    elif vet_score >= 1:
        score += vet_score
        breakdown['veteran'] = vet_score
        reasons.append(f"🎖️ Experience ({white_data['total_games']}+{black_data['total_games']} games)")
    
    # RETURN MATCH (0-4 points) - tiered by length of absence
    for player, data in [(white, white_data), (black, black_data)]:
        seasons = data.get('seasons_played', set())
        games_per_season = data.get('games_per_season', {})
        
        if not seasons:
            continue
            
        # Check if they have any games this season already (before this round)
        games_this_season = games_per_season.get(current_season, 0)
        if games_this_season > 0:
            # Already played this season, not their first game back
            continue
        
        # Find their most recent season before current
        past_seasons = sorted([s for s in seasons if s < current_season])
        if not past_seasons:
            continue  # No history, this is their debut
        
        last_played = past_seasons[-1]
        gap = current_season - last_played
        
        if gap >= 20:
            pts = 4
            score += pts
            breakdown['return_match'] += pts
            reasons.append(f"🪃 {player}'s epic return after {gap} seasons!")
        elif gap >= 10:
            pts = 3
            score += pts
            breakdown['return_match'] += pts
            reasons.append(f"🪃 {player}'s big return after {gap} seasons!")
        elif gap >= 5:
            pts = 2
            score += pts
            breakdown['return_match'] += pts
            reasons.append(f"🪃 {player}'s return after {gap} seasons!")
    
    # FORMER TEAMMATES (0-5 points) - both played 4+ games on same team(s)
    # +1 bonus if they won the league together
    white_teams = white_data.get('teams', {})
    black_teams = black_data.get('teams', {})
    white_champs = white_data.get('championship_teams', set())
    black_champs = black_data.get('championship_teams', set())
    
    # Find teams where BOTH played 4+ games
    shared_teams = []
    for team, white_games in white_teams.items():
        if white_games >= 4 and black_teams.get(team, 0) >= 4:
            shared_teams.append(team)
    
    # Check if any shared team is a championship team for both
    shared_championship = any(team in white_champs and team in black_champs for team in shared_teams)
    
    if len(shared_teams) >= 3:
        pts = 4
        score += pts
        breakdown['former_teammates'] = pts
        if shared_championship:
            pts_bonus = 1
            score += pts_bonus
            breakdown['former_teammates'] += pts_bonus
            reasons.append(f"🤝🏆 Old pals & league champs! ({len(shared_teams)}x teammates)")
        else:
            reasons.append(f"🤝 Old pals! ({len(shared_teams)}x teammates)")
    elif len(shared_teams) == 2:
        pts = 3
        score += pts
        breakdown['former_teammates'] = pts
        if shared_championship:
            pts_bonus = 1
            score += pts_bonus
            breakdown['former_teammates'] += pts_bonus
            reasons.append(f"🤝🏆 Twice teammates & league champs!")
        else:
            reasons.append(f"🤝 Twice teammates! ({shared_teams[0]}, {shared_teams[1]})")
    elif len(shared_teams) == 1:
        pts = 2
        score += pts
        breakdown['former_teammates'] = pts
        if shared_championship:
            pts_bonus = 1
            score += pts_bonus
            breakdown['former_teammates'] += pts_bonus
            reasons.append(f"🤝🏆 Former teammates & league champs on {shared_teams[0]}")
        else:
            reasons.append(f"🤝 Former teammates on {shared_teams[0]}")
    
    # VETS FIRST MEETING (0-1.5 points) - experienced players who never faced each other
    white_games = white_data.get('total_games', 0)
    black_games = black_data.get('total_games', 0)
    never_played = h2h_games == 0
    
    if never_played:
        if white_games >= 100 and black_games >= 100:
            pts = 1.5
            score += pts
            breakdown['vets_first_meeting'] = pts
            reasons.append(f"👀 Vets first meeting! ({white_games}+{black_games} games, never faced)")
        elif white_games >= 80 and black_games >= 80:
            pts = 0.75
            score += pts
            breakdown['vets_first_meeting'] = pts
            reasons.append(f"👀 Vets first meeting! ({white_games}+{black_games} games, never faced)")
    
    # DEBUTANTS' DUEL (1.5 points) - both players making their league debut
    white_games = white_data.get('total_games', 0)
    black_games = black_data.get('total_games', 0)
    if white_games == 0 and black_games == 0:
        pts = 1.5
        score += pts
        breakdown['debutants_duel'] = pts
        reasons.append(f"🌱 Debutants' duel! Both making their league debut")
    
    return score, reasons, breakdown


def rank_team_battles(pairings, standings):
    """Rank team matchups by average team placement."""
    team_standings = {t['name']: i + 1 for i, t in enumerate(standings)}
    
    # Get unique team matchups
    matchups = {}
    for p in pairings:
        white_team = p.get('white_team') or p.get('white_team_name', '')
        black_team = p.get('black_team') or p.get('black_team_name', '')
        
        if not white_team or not black_team:
            continue
        
        match_key = tuple(sorted([white_team, black_team]))
        if match_key not in matchups:
            white_rank = team_standings.get(white_team, 99)
            black_rank = team_standings.get(black_team, 99)
            avg_rank = (white_rank + black_rank) / 2
            matchups[match_key] = avg_rank
    
    # Sort by average rank (lower = better teams)
    sorted_matchups = sorted(matchups.items(), key=lambda x: x[1])
    
    # Assign ranks
    ranks = {}
    for i, (match_key, _) in enumerate(sorted_matchups):
        ranks[match_key] = i + 1
    
    return ranks


def filter_season_by_round(season_data, before_round):
    """Return season data with only decided games before the target round."""
    filtered = dict(season_data)
    filtered['games'] = [
        g for g in season_data.get('games', [])
        if g.get('round', 0) < before_round and g.get('result', '') != ''  # Has result before target round
    ]
    return filtered


def generate_html_output(scored_pairings, season, round_num, output_file):
    """Generate HTML output with teal color scheme."""
    
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Storylines - Season {season} Round {round_num}</title>
    <style>
        :root {{
            --bg-primary: #2d5266;
            --bg-dark: #234250;
            --bg-darker: #1a323d;
            --bg-card: rgba(255, 255, 255, 0.05);
            --text-primary: #f0f6f6;
            --text-secondary: rgba(240, 246, 246, 0.7);
            --text-muted: rgba(240, 246, 246, 0.4);
            --border: rgba(255, 255, 255, 0.12);
            --gold: #d4b856;
            --accent: #4a9ebb;
        }}
        
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, var(--bg-darker) 0%, var(--bg-dark) 50%, var(--bg-primary) 100%);
            color: var(--text-primary);
            min-height: 100vh;
            padding: 2rem;
        }}
        
        .container {{
            max-width: 900px;
            margin: 0 auto;
        }}
        
        header {{
            text-align: center;
            margin-bottom: 2rem;
            padding: 2rem;
            background: var(--bg-card);
            border: 1px solid var(--border);
            border-radius: 12px;
        }}
        
        header h1 {{
            font-size: 2rem;
            margin-bottom: 0.5rem;
            color: var(--text-primary);
        }}
        
        header .subtitle {{
            color: var(--text-secondary);
            font-size: 1.1rem;
        }}
        
        header .stats {{
            margin-top: 1rem;
            color: var(--text-muted);
            font-size: 0.9rem;
        }}
        
        .pairing-card {{
            background: var(--bg-card);
            border: 1px solid var(--border);
            border-radius: 12px;
            padding: 1.5rem;
            margin-bottom: 1rem;
            transition: transform 0.2s, border-color 0.2s;
        }}
        
        .pairing-card:hover {{
            transform: translateY(-2px);
            border-color: var(--accent);
        }}
        
        .pairing-header {{
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            margin-bottom: 1rem;
        }}
        
        .rank {{
            font-size: 1.5rem;
            font-weight: bold;
            color: var(--gold);
            min-width: 50px;
        }}
        
        .rank.top-3 {{
            font-size: 1.8rem;
        }}
        
        .matchup {{
            flex-grow: 1;
            text-align: center;
        }}
        
        .players {{
            font-size: 1.3rem;
            font-weight: 600;
            margin-bottom: 0.3rem;
        }}
        
        .players a {{
            color: var(--text-primary);
            text-decoration: none;
        }}
        
        .players a:hover {{
            color: var(--accent);
            text-decoration: underline;
        }}
        
        .vs {{
            color: var(--text-muted);
            margin: 0 0.5rem;
        }}
        
        .teams {{
            color: var(--text-secondary);
            font-size: 0.9rem;
        }}
        
        .schedule-time {{
            font-size: 0.85rem;
            margin-top: 0.4rem;
            padding: 0.2rem 0.5rem;
            border-radius: 4px;
            display: inline-block;
        }}
        
        .schedule-time.scheduled {{
            color: #4ade80;
            background: rgba(74, 222, 128, 0.1);
        }}
        
        .schedule-time.tbd {{
            color: var(--text-muted);
        }}
        
        .schedule-time.played {{
            color: #60a5fa;
            background: rgba(96, 165, 250, 0.1);
        }}
        
        .score-badge {{
            background: var(--bg-dark);
            border: 1px solid var(--border);
            border-radius: 8px;
            padding: 0.5rem 1rem;
            text-align: center;
            min-width: 70px;
        }}
        
        .score-value {{
            font-size: 1.2rem;
            font-weight: bold;
            color: var(--gold);
        }}
        
        .score-label {{
            font-size: 0.7rem;
            color: var(--text-muted);
            text-transform: uppercase;
        }}
        
        .reasons {{
            margin-top: 1rem;
            padding-top: 1rem;
            border-top: 1px solid var(--border);
        }}
        
        .reason {{
            display: inline-block;
            background: rgba(255, 255, 255, 0.06);
            border-radius: 6px;
            padding: 0.4rem 0.8rem;
            margin: 0.2rem;
            font-size: 0.9rem;
            color: var(--text-secondary);
        }}
        
        footer {{
            text-align: center;
            margin-top: 2rem;
            padding: 1rem;
            color: var(--text-muted);
            font-size: 0.85rem;
        }}
        
        footer a {{
            color: var(--accent);
            text-decoration: none;
        }}
        
        .guide-toggle {{
            background: var(--accent);
            color: var(--text-primary);
            padding: 0.5rem 1rem;
            border-radius: 6px;
            cursor: pointer;
            font-size: 0.9rem;
            margin-top: 1rem;
            transition: background 0.2s;
            display: inline-block;
            list-style: none;
        }}
        
        .guide-toggle::-webkit-details-marker {{
            display: none;
        }}
        
        .guide-toggle:hover {{
            background: #5ab0cc;
        }}
        
        .guide-wrapper {{
            margin-bottom: 2rem;
        }}
        
        .guide-wrapper[open] .guide-toggle {{
            margin-bottom: 1rem;
        }}
        
        .guide {{
            background: var(--bg-card);
            border: 1px solid var(--border);
            border-radius: 12px;
            padding: 1.5rem;
            text-align: left;
        }}
        
        .guide h2 {{
            font-size: 1.3rem;
            margin-bottom: 1rem;
            color: var(--gold);
        }}
        
        .guide h3 {{
            font-size: 1rem;
            margin-top: 1rem;
            margin-bottom: 0.5rem;
            color: var(--text-primary);
        }}
        
        .guide table {{
            width: 100%;
            border-collapse: collapse;
            margin: 0.5rem 0;
        }}
        
        .guide th, .guide td {{
            padding: 0.5rem;
            text-align: left;
            border-bottom: 1px solid var(--border);
        }}
        
        .guide th {{
            color: var(--text-secondary);
            font-weight: normal;
            font-size: 0.85rem;
        }}
        
        .guide td {{
            color: var(--text-primary);
            font-size: 0.9rem;
        }}
        
        .guide .points {{
            color: var(--gold);
            font-weight: bold;
        }}
        
        @media (max-width: 600px) {{
            body {{
                padding: 1rem;
            }}
            
            .pairing-header {{
                flex-direction: column;
                align-items: center;
                gap: 1rem;
            }}
            
            .rank {{
                order: -1;
            }}
            
            .players {{
                font-size: 1.1rem;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>📺 Storylines of the Week</h1>
            <div class="subtitle">Season {season} &middot; Round {round_num}</div>
            <div class="stats">Top 10 of {len(scored_pairings)} pairings</div>
        </header>
        
        <details class="guide-wrapper">
            <summary class="guide-toggle">📊 How is Hype Score calculated?</summary>
            <div class="guide">
            <h2>📊 Hype Score Calculation</h2>
            <p style="color: var(--text-secondary); margin-bottom: 1rem;">
                Each pairing is scored based on several factors that make a matchup compelling:
            </p>
            
            <h3>👑 Titled Player (0-3 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>One titled player (FM, IM, GM, etc.)</td><td class="points">+1.5</td></tr>
                <tr><td>Both players titled</td><td class="points">+3</td></tr>
            </table>
            
            <h3>🎯 Milestone Games (0-6 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>Player's 45th career game</td><td class="points">+2</td></tr>
                <tr><td>Player's 100th career game</td><td class="points">+3</td></tr>
                <tr><td>Player's 200th/300th career game</td><td class="points">+3</td></tr>
            </table>
            <p style="color: var(--text-muted); font-size: 0.85rem;">Can stack if both players have milestones</p>
            
            <h3>🏆 100th Win (0-2.5 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>Player at 99 wins (going for 100th)</td><td class="points">+2.5</td></tr>
            </table>
            
            <h3>⚔️ Head-to-Head History (0-4 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>Played once before</td><td class="points">+2</td></tr>
                <tr><td>Played twice before</td><td class="points">+3</td></tr>
                <tr><td>Played 3+ times (Rivals!)</td><td class="points">+4</td></tr>
            </table>
            
            <h3>🔥 Form Clash (0-3 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>Both on 2+ win streak</td><td class="points">+2</td></tr>
                <tr><td>Both on 3+ win streak</td><td class="points">+3</td></tr>
                <tr><td>Both on 2+ loss streak</td><td class="points">+2</td></tr>
                <tr><td>Both on 3+ loss streak</td><td class="points">+3</td></tr>
            </table>
            <p style="color: var(--text-muted); font-size: 0.85rem;">Streaks are career-wide, not just current season</p>
            
            <h3>🏆 Team Standings Impact (0-6 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Base Points</th></tr>
                <tr><td>#1 team matchup (by avg placement)</td><td class="points">+4</td></tr>
                <tr><td>#2 team matchup</td><td class="points">+3</td></tr>
                <tr><td>#3 team matchup</td><td class="points">+2</td></tr>
            </table>
            <p style="color: var(--text-muted); font-size: 0.85rem; margin-top: 0.5rem;">
                <strong>Round multiplier:</strong> R1: 0x, R2: 0.25x, R3: 0.5x, R4: 0.75x, R5-6: 1x, R7-8: 1.5x
            </p>
            
            <h3>🎖️ Veteran Experience (0-4 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>Career games / 100 (per player, max 2 each)</td><td class="points">0-4</td></tr>
            </table>
            <p style="color: var(--text-muted); font-size: 0.85rem;">e.g., 150 games = 1.5 pts, 200+ games = 2 pts (capped)</p>
            
            <h3>🪃 Return Match (0-4 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>First game back after 5-9 seasons</td><td class="points">+2</td></tr>
                <tr><td>First game back after 10-19 seasons</td><td class="points">+3</td></tr>
                <tr><td>First game back after 20+ seasons</td><td class="points">+4</td></tr>
            </table>
            
            <h3>🤝 Former Teammates (0-5 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>Both played 4+ games on 1 shared team</td><td class="points">+2</td></tr>
                <tr><td>Both played 4+ games on 2 shared teams</td><td class="points">+3</td></tr>
                <tr><td>Both played 4+ games on 3+ shared teams</td><td class="points">+4</td></tr>
                <tr><td>Bonus: Won league together</td><td class="points">+1</td></tr>
            </table>
            
            <h3>👀 Vets First Meeting (0-1.5 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>Both 80+ games, never faced each other</td><td class="points">+0.75</td></tr>
                <tr><td>Both 100+ games, never faced each other</td><td class="points">+1.5</td></tr>
            </table>
            
            <h3>🌱 Debutants' Duel (1.5 pts)</h3>
            <table>
                <tr><th>Condition</th><th>Points</th></tr>
                <tr><td>Both players making their league debut</td><td class="points">+1.5</td></tr>
            </table>
            
            <p style="color: var(--text-secondary); margin-top: 1.5rem;">
                <strong>Maximum possible score: ~38.5 points</strong>
            </p>
            </div>
        </details>
        
        <div class="pairings">
"""
    
    for i, p in enumerate(scored_pairings[:10], 1):
        rank_class = "top-3" if i <= 3 else ""
        
        # Build Lichess profile links
        white_link = f"https://www.lichess4545.com/team4545/player/{p['white']}/"
        black_link = f"https://www.lichess4545.com/team4545/player/{p['black']}/"
        
        # Build schedule status
        if p.get('played'):
            schedule_html = f'<div class="schedule-time played">✓ {p.get("result", "played")}</div>'
        elif p.get('scheduled_time'):
            time_str = p['scheduled_time'].strftime('%a %b %d, %H:%M UTC')
            schedule_html = f'<div class="schedule-time scheduled">🕐 {time_str}</div>'
        else:
            schedule_html = '<div class="schedule-time tbd">⏳ TBD</div>'
        
        html += f"""
            <div class="pairing-card">
                <div class="pairing-header">
                    <div class="rank {rank_class}">#{i}</div>
                    <div class="matchup">
                        <div class="players">
                            <a href="{white_link}" target="_blank">{p['white']}</a>
                            <span class="vs">vs</span>
                            <a href="{black_link}" target="_blank">{p['black']}</a>
                        </div>
                        <div class="teams">{p['white_team']} vs {p['black_team']}</div>
                        {schedule_html}
                    </div>
                    <div class="score-badge">
                        <div class="score-value">{p['score']:.1f}</div>
                        <div class="score-label">hype</div>
                    </div>
                </div>
"""
        
        if p['reasons']:
            html += '                <div class="reasons">\n'
            for reason in p['reasons']:
                html += f'                    <span class="reason">{reason}</span>\n'
            html += '                </div>\n'
        
        html += '            </div>\n'
    
    html += f"""
        </div>
        
        <footer>
            Generated for <a href="https://www.lichess4545.com" target="_blank">Lichess4545</a> League
        </footer>
    </div>
</body>
</html>
"""
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"\n✓ HTML output saved to: {output_file}")


def generate_storylines(season, target_round=None, output_html=None, output_excel=None, no_refresh=False, ics_path=None):
    """Generate storylines for upcoming round."""
    
    # Auto-refresh current season data (unless --no-refresh)
    if not no_refresh:
        refresh_season_cache(season, force=False)
    
    print(f"Loading season {season} data...")
    season_data = load_season(season)
    if not season_data:
        return
    
    # Find the target round first
    print("Finding round pairings...")
    round_num, pairings = get_round_pairings(season_data, target_round)
    
    if not pairings:
        print("No upcoming pairings found!")
        return
    
    print(f"Found {len(pairings)} pairings for Round {round_num}")
    print(f"(Using data from rounds 1-{round_num - 1} for calculations)")
    
    # Parse ICS for scheduled times if provided
    schedule_lookup = {}
    if ics_path:
        print(f"\nLoading scheduled times from {ics_path}...")
        ics_events = parse_ics_file(ics_path)
        schedule_lookup = build_schedule_lookup(ics_events)
        print(f"  Found {len(schedule_lookup)} scheduled games")
    
    print(f"\nLoading historical data (seasons 1-{season})...")
    all_seasons = load_all_seasons(season)
    
    print("Identifying league champions...")
    championship_teams = get_championship_teams(all_seasons)
    print(f"  Found {len(championship_teams)} completed seasons with champions")
    
    print("Building player history (filtered to before target round)...")
    player_history = build_player_history(all_seasons, current_season=season, before_round=round_num, championship_teams=championship_teams)
    
    print("Calculating current standings (before target round)...")
    filtered_season = filter_season_by_round(season_data, round_num)
    standings = calculate_standings(filtered_season)
    
    print("Analyzing career streaks (before target round)...")
    streaks = get_career_streaks(all_seasons, season, round_num)
    
    # Rank team battles
    team_battle_ranks = rank_team_battles(pairings, standings)
    
    # Fetch player titles from Lichess API
    all_players = set()
    for p in pairings:
        if p.get('white'):
            all_players.add(p['white'].lower())
        if p.get('black'):
            all_players.add(p['black'].lower())
    player_titles = fetch_player_titles(list(all_players))
    
    # Calculate hype scores
    scored_pairings = []
    for pairing in pairings:
        score, reasons, breakdown = calculate_hype_score(
            pairing, player_history, standings, streaks, team_battle_ranks, season, player_titles, round_num
        )
        
        # Look up scheduled time from ICS
        white = pairing.get('white', '').lower()
        black = pairing.get('black', '').lower()
        player_key = tuple(sorted([white, black]))
        scheduled_time = schedule_lookup.get(player_key)
        
        scored_pairings.append({
            'white': pairing.get('white', ''),
            'black': pairing.get('black', ''),
            'white_team': pairing.get('white_team') or pairing.get('white_team_name', ''),
            'black_team': pairing.get('black_team') or pairing.get('black_team_name', ''),
            'score': score,
            'reasons': reasons,
            'breakdown': breakdown,
            'played': bool(pairing.get('game_id')),
            'result': pairing.get('result', ''),
            'scheduled_time': scheduled_time,
        })
    
    # Sort by hype score
    scored_pairings.sort(key=lambda x: -x['score'])
    
    # Count played/unplayed/scheduled
    played_count = sum(1 for p in scored_pairings if p['played'])
    unplayed_count = len(scored_pairings) - played_count
    scheduled_count = sum(1 for p in scored_pairings if p.get('scheduled_time') and not p['played'])
    
    # Print results
    print(f"\n{'='*70}")
    print(f"  📺 STORYLINES OF THE WEEK - Season {season} Round {round_num}")
    print(f"{'='*70}")
    if schedule_lookup:
        print(f"  ({played_count} played, {scheduled_count} scheduled, {unplayed_count - scheduled_count} TBD)\n")
    else:
        print(f"  ({played_count} played, {unplayed_count} unplayed)\n")
    
    for i, p in enumerate(scored_pairings[:15], 1):
        if p['played']:
            status = f"[{p['result']}]"
        elif p.get('scheduled_time'):
            status = f"[{p['scheduled_time'].strftime('%a %b %d %H:%M')} UTC]"
        else:
            status = "[TBD]"
        print(f"#{i} [{p['score']:.1f} pts] {p['white']} vs {p['black']} {status}")
        print(f"   ({p['white_team']} vs {p['black_team']})")
        for reason in p['reasons']:
            print(f"   • {reason}")
        print()
    
    if len(scored_pairings) > 15:
        print(f"... and {len(scored_pairings) - 15} more pairings")
    
    # Generate HTML if requested
    if output_html:
        generate_html_output(scored_pairings, season, round_num, output_html)
    
    # Generate Excel if requested
    if output_excel:
        generate_excel_output(scored_pairings, season, round_num, output_excel)
    
    return scored_pairings


def generate_excel_output(scored_pairings, season, round_num, output_file):
    """Generate Excel file with detailed breakdown of all pairings."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Warning: openpyxl not installed. Run: pip install openpyxl")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"S{season} R{round_num} Hype"
    
    # Define breakdown columns
    breakdown_cols = [
        ('titled', 'Titled'),
        ('milestone', 'Milestone'),
        ('win_milestone', '100th Win'),
        ('h2h', 'H2H'),
        ('form_clash', 'Form'),
        ('team_battle', 'Team Battle'),
        ('veteran', 'Veteran'),
        ('return_match', 'Return'),
        ('former_teammates', 'Teammates'),
        ('vets_first_meeting', 'Vets 1st'),
        ('debutants_duel', 'Deb Duel'),
    ]
    
    # Headers
    headers = ['Rank', 'White', 'Black', 'White Team', 'Black Team', 'Total', 'Status']
    headers += [col[1] for col in breakdown_cols]
    headers += ['Reasons', 'Team Matchup']
    
    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2D5A5A')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Write data rows
    for rank, pairing in enumerate(scored_pairings, 1):
        row = rank + 1
        breakdown = pairing.get('breakdown', {})
        
        # Basic info
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=pairing['white'])
        ws.cell(row=row, column=3, value=pairing['black'])
        ws.cell(row=row, column=4, value=pairing['white_team'])
        ws.cell(row=row, column=5, value=pairing['black_team'])
        ws.cell(row=row, column=6, value=round(pairing['score'], 1))
        ws.cell(row=row, column=7, value=pairing['result'] if pairing['played'] else 'upcoming')
        
        # Individual breakdown columns
        for col_idx, (key, _) in enumerate(breakdown_cols, 8):
            val = breakdown.get(key, 0)
            cell = ws.cell(row=row, column=col_idx, value=val if val else '')
        
        # Reasons
        reasons_col = 8 + len(breakdown_cols)
        reasons_text = '; '.join(pairing.get('reasons', []))
        ws.cell(row=row, column=reasons_col, value=reasons_text)
        
        # Team Matchup (for filtering)
        matchup_col = reasons_col + 1
        team_matchup = ' vs '.join(sorted([pairing['white_team'], pairing['black_team']]))
        ws.cell(row=row, column=matchup_col, value=team_matchup)
        
        # Apply borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
        
        # Highlight top 10 pairings
        if rank <= 10:
            highlight_fill = PatternFill('solid', fgColor='E8F5F5')
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = highlight_fill
    
    # Adjust column widths
    col_widths = {
        1: 6,   # Rank
        2: 18,  # White
        3: 18,  # Black
        4: 20,  # White Team
        5: 20,  # Black Team
        6: 7,   # Total
        7: 10,  # Status
    }
    # Breakdown columns
    for i in range(8, 8 + len(breakdown_cols)):
        col_widths[i] = 10
    # Reasons column
    col_widths[8 + len(breakdown_cols)] = 60
    # Team Matchup column
    col_widths[9 + len(breakdown_cols)] = 30
    
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary sheet
    summary = wb.create_sheet('Summary')
    summary['A1'] = f"Season {season} Round {round_num} Hype Score Analysis"
    summary['A1'].font = Font(bold=True, size=14)
    
    summary['A3'] = "Factor Distribution"
    summary['A3'].font = Font(bold=True)
    
    # Count how many pairings got each factor
    row = 4
    for key, display in breakdown_cols:
        count = sum(1 for p in scored_pairings if p.get('breakdown', {}).get(key, 0) > 0)
        total_pts = sum(p.get('breakdown', {}).get(key, 0) for p in scored_pairings)
        summary.cell(row=row, column=1, value=display)
        summary.cell(row=row, column=2, value=count)
        summary.cell(row=row, column=3, value=f"({total_pts:.1f} pts)")
        row += 1
    
    summary['A15'] = "Score Distribution"
    summary['A15'].font = Font(bold=True)
    
    # Score buckets
    buckets = {'0': 0, '1-3': 0, '4-6': 0, '7-9': 0, '10+': 0}
    for p in scored_pairings:
        s = p['score']
        if s == 0:
            buckets['0'] += 1
        elif s <= 3:
            buckets['1-3'] += 1
        elif s <= 6:
            buckets['4-6'] += 1
        elif s <= 9:
            buckets['7-9'] += 1
        else:
            buckets['10+'] += 1
    
    row = 16
    for bucket, count in buckets.items():
        summary.cell(row=row, column=1, value=bucket)
        summary.cell(row=row, column=2, value=count)
        row += 1
    
    # Add Team Matchups sheet
    team_sheet = wb.create_sheet('Team Matchups')
    team_sheet['A1'] = f"Season {season} Round {round_num} - Team Matchup Hype Scores"
    team_sheet['A1'].font = Font(bold=True, size=14)
    
    # Group pairings by team matchup
    team_matchups = defaultdict(lambda: {'total_score': 0, 'games': [], 'game_count': 0})
    
    for p in scored_pairings:
        white_team = p.get('white_team', '')
        black_team = p.get('black_team', '')
        if white_team and black_team:
            # Sort team names for consistent key
            matchup_key = ' vs '.join(sorted([white_team, black_team]))
            team_matchups[matchup_key]['total_score'] += p['score']
            team_matchups[matchup_key]['game_count'] += 1
            team_matchups[matchup_key]['games'].append(p)
    
    # Sort by total score descending
    sorted_matchups = sorted(team_matchups.items(), key=lambda x: -x[1]['total_score'])
    
    # Headers
    headers = ['Rank', 'Team Matchup', 'Total Hype', 'Games', 'Avg/Game', 'Top Pairing']
    for col, header in enumerate(headers, 1):
        cell = team_sheet.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2D5A5A')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for rank, (matchup, data) in enumerate(sorted_matchups, 1):
        row = rank + 3
        avg_score = data['total_score'] / data['game_count'] if data['game_count'] > 0 else 0
        
        # Find top pairing in this matchup
        top_game = max(data['games'], key=lambda x: x['score'])
        top_pairing = f"{top_game['white']} vs {top_game['black']} ({top_game['score']:.1f})"
        
        team_sheet.cell(row=row, column=1, value=rank)
        team_sheet.cell(row=row, column=2, value=matchup)
        team_sheet.cell(row=row, column=3, value=round(data['total_score'], 1))
        team_sheet.cell(row=row, column=4, value=data['game_count'])
        team_sheet.cell(row=row, column=5, value=round(avg_score, 2))
        team_sheet.cell(row=row, column=6, value=top_pairing)
        
        # Highlight top 3
        if rank <= 3:
            for col in range(1, 7):
                team_sheet.cell(row=row, column=col).fill = PatternFill('solid', fgColor='E8F5F5')
    
    # Column widths
    team_sheet.column_dimensions['A'].width = 6
    team_sheet.column_dimensions['B'].width = 40
    team_sheet.column_dimensions['C'].width = 12
    team_sheet.column_dimensions['D'].width = 8
    team_sheet.column_dimensions['E'].width = 10
    team_sheet.column_dimensions['F'].width = 35
    
    team_sheet.freeze_panes = 'A4'
    
    # Save
    wb.save(output_file)
    print(f"\nExcel output saved to {output_file}")


def main():
    parser = argparse.ArgumentParser(
        description='Generate storylines for upcoming Lichess4545 pairings'
    )
    parser.add_argument('--season', '-s', type=int, required=True,
                        help='Season number')
    parser.add_argument('--round', '-r', type=int, default=None,
                        help='Round number (auto-detect if not specified)')
    parser.add_argument('--output', '-o', type=str, default=None,
                        help='Output HTML file (e.g., storylines.html)')
    parser.add_argument('--excel', '-e', type=str, default=None,
                        help='Output Excel file with detailed breakdown (e.g., storylines.xlsx)')
    parser.add_argument('--ics', type=str, default=None,
                        help='ICS calendar file for scheduled game times')
    parser.add_argument('--no-refresh', action='store_true',
                        help='Skip auto-refresh of season data (use cached data)')
    parser.add_argument('--json', action='store_true',
                        help='Output as JSON')
    
    args = parser.parse_args()
    
    result = generate_storylines(args.season, args.round, args.output, args.excel, args.no_refresh, args.ics)
    
    if args.json and result:
        print(json.dumps(result, indent=2, default=str))


if __name__ == '__main__':
    main()
